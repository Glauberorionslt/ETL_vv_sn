import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime

# Configuração do logger
log_file = "meu_log.log"
logging.basicConfig(filename=log_file, level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def combinar_arquivos():
    pasta = r"Z:\#BASES DE APOIO - GERENCIAL\01 - BASES\PROJETOS\SERVICE NOW"
    dfs = []

    for filename in os.listdir(pasta):
        if filename.endswith(".xlsx") and not filename.startswith('~$'):
            arquivo = os.path.join(pasta, filename)
            
            try:
                df = pd.read_excel(arquivo, sheet_name="Page 1")
                dfs.append(df)
                logging.info(f"Arquivo '{filename}' lido com sucesso.")
            
            except Exception as e:
                logging.error(f"Erro ao ler o arquivo '{filename}': {e}")

    if dfs:
        df_final = pd.concat(dfs, ignore_index=True)
        default_font = Font(name='Calibri', size=12)
        wb = Workbook()
        ws = wb.active
        
        for r in dataframe_to_rows(df_final, index=False, header=True):
            ws.append(r)
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font
        
        output_directory = r"C:\Users\2160011883.VIA\OneDrive - Via Varejo S.A\GLAUBER S-ONE DRIVE\PROJETOS\PYTHON\Analise de dados\ETL\Pbi\indicadores_manutencao"
        output_file = os.path.join(output_directory, "f_Servicenow_combinado.xlsx")
        
        if os.path.exists(output_file):
            os.remove(output_file)
            logging.info(f"Arquivo existente '{output_file}' removido.")
        
        wb.save(output_file)
        logging.info(f"Arquivo combinado salvo em: {output_file}")
    else:
        logging.warning("Nenhum arquivo válido encontrado para combinar.")


def preprocessar_servicenow_file(input_file):
    def processar_dados(df):
        def verificar_estado(estado):
            if estado in ['Resolvido', 'Encerrado']:
                return 'ENCERRADO'
            else:
                return 'ABERTO'

        def verificar_prioridade(prioridade):
            if prioridade == '1 - Crítico':
                return 'P1'
            elif prioridade == '2 - Alto':
                return 'P2'
            elif prioridade == '3 - Moderado':
                return 'P3'
            elif prioridade == '4 - Baixo':
                return 'P4'
            else:
                return None

        def verificar_codigo_solucao(codigo_solucao):
            if codigo_solucao in ['Solucionado com Custo', 'Executado sem custo', 'Solucionado (Reprovado)',
                                   'Solicitação Atendida']:
                return 'Solucionado'
            elif codigo_solucao == 'Fora do Escopo':
                return 'Escopo não Aprovado'
            elif codigo_solucao in ['Duplicidade', 'Atendido em outro chamado']:
                return 'Abertura incorreta'
            else:
                return codigo_solucao

        df['Status - Estado'] = df['Estado'].apply(verificar_estado)
        df['PRIORIDADE RES'] = df['Prioridade'].apply(verificar_prioridade)
        df['Cod ajustado'] = df['Código de Solução'].apply(verificar_codigo_solucao)

        return df

    df = pd.read_excel(input_file)
    colunas_desejadas = [
        "Item", "Número", "Item.1", "Código", "Aberto", "Prioridade", "Estado", "Descrição resumida", 
        "Descrição", "Grupo designado", "Atribuído a", "Código de fechamento", "Código de Solução", 
        "Resolução", "Resolvido", "Encerrado", " Nome do Fornecedor Acionado", "Valor Total", 
        "Número OM", "Número OI ", "Número do Pedido (Ariba) ", "Número da Requisição de Compras"
    ]
    colunas_presentes = [col for col in colunas_desejadas if col in df.columns]
    df_reduzido = df[colunas_presentes]

    df_processado = processar_dados(df_reduzido)
    logging.info("DataFrame processado com sucesso.")

    output_path = r"C:\Users\2160011883.VIA\OneDrive - Via Varejo S.A\Bases - DB_MANUTENCAO"
    output_file = os.path.join(output_path, "f_Servicenow.xlsx")

    if os.path.exists(output_file):
        os.remove(output_file)
        logging.info(f"Arquivo existente '{output_file}' removido.")

    df_processado.to_excel(output_file, index=False)
    logging.info(f"DataFrame processado salvo com sucesso em '{output_file}'")


if __name__ == "__main__":
    # Executar a função combinar_arquivos()
 combinar_arquivos()

    # Exemplo de uso da função preprocessar_servicenow_file com o arquivo combinado
input_file = 'f_Servicenow_combinado.xlsx'
preprocessar_servicenow_file(input_file)



# # Nome do arquivo a ser verificado (incluindo extensão)
# nome_arquivo = "f_Servicenow_combinado.xlsx"

# # Verificar se o arquivo existe no diretório atual
# if os.path.exists(nome_arquivo):
#     print(f"O arquivo '{nome_arquivo}' EXISTE no diretório atual.")
# else:
#     print(f"O arquivo '{nome_arquivo}' não existe no diretório atual.")